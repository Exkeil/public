import psycopg2
import pandas as pd
import openpyxl
from openpyxl.utils import get_column_letter
from datetime import datetime
from db_config import conn_params

table_name = 'patriots'

with psycopg2.connect(**conn_params) as conn, conn.cursor() as cursor:
    query = f'''
    SELECT 
        employee_id AS "Сотрудник",
        id AS "№",
        export_number AS "Выгрузка 1",
        candidate_id AS "ID",
        previous_stage AS "Предыдущая стадия",
        full_name AS "Название",
        refusal_reason_new AS "Причина отказа актуал",
        contact AS "Контакт",
        phone AS "Номер телефона",
        age AS "Возраст",
        city AS "Город",
        stage_new AS "Стадия",
        refusal_reason AS "Причина отказа",
        comment AS "Комментарий"
    FROM {table_name}
    WHERE 
        stage_new NOT IN ('Отказ', 'Не набран 1 балл')
        AND (refusal_reason IS NULL OR refusal_reason NOT IN (
            'Старше 21 года',
            'Номер не существует',
            '8 класс',
            '7 класс и ниже',
            'Трудоустройство',
            'Попросил никогда больше не беспокоить',
            'Сотрудник',
            'Недозвон СНГ'
        ))
    LIMIT 25000;
    '''
    cursor.execute(query)
    df = pd.DataFrame(cursor.fetchall(), columns=[desc[0] for desc in cursor.description])
    df = df.fillna('')
    df['Возраст'] = pd.to_numeric(df['Возраст'], errors='coerce').fillna(0).astype(int)
    df['Номер телефона'] = df['Номер телефона'].astype(str)
    output_file = f"Итоговая_выгрузка_{datetime.now().strftime('%d_%m_%Y')}.xlsx"
    df.to_excel(output_file, sheet_name='Sheet1', index=False, engine='openpyxl')
    wb = openpyxl.load_workbook(output_file)
    ws = wb['Sheet1']
    visible_columns = [
        'Сотрудник', '№', 'Контакт', 'Номер телефона', 'Возраст',
        'Город', 'Стадия', 'Причина отказа', 'Комментарий'
    ]
    for col_num, col_name in enumerate(df.columns, 1):
        max_len = max(df[col_name].astype(str).map(len).max(), len(col_name)) + 2
        col_letter = get_column_letter(col_num)
        ws.column_dimensions[col_letter].width = max_len
        if col_name not in visible_columns:
            ws.column_dimensions[col_letter].hidden = True
    wb.save(output_file)
    print(f"Файл успешно экспортирован: {output_file}")
    print(f"Экспортировано строк: {len(df)}")
